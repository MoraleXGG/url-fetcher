"""Carga URLs desde archivos CSV, JSON o TXT."""

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

# Nombres de columna URL que se buscan automáticamente, en orden de preferencia.
# El matching es case-insensitive; las variantes que solo difieren en case
# están aquí como pista de los formatos típicos para futuros mantenedores.
DEFAULT_URL_COLUMNS = [
    "url",                # genérico
    "URL",                # Search Console y otros
    "Original Url",       # Screaming Frog (export estándar)
    "Address",            # Screaming Frog (otras vistas)
    "loc",                # sitemap.xml
    "link",
    "Página",             # Search Console español
    "Top URL",            # Search Console: páginas top
    "Página principal",   # Search Console español alternativo
]


def load_urls(path: Path, url_column: str | None = None) -> list[str]:
    """Carga URLs despachando por extensión (.txt/.csv/.json).

    Lanza:
        FileNotFoundError: si `path` no existe.
        ValueError: si la extensión no es soportada o no se puede detectar la columna.
    """
    if not path.exists():
        raise FileNotFoundError(f"Archivo no encontrado: {path}")

    suffix = path.suffix.lower()
    if suffix == ".txt":
        return _load_txt(path)
    if suffix == ".csv":
        return _load_csv(path, url_column)
    if suffix == ".json":
        return _load_json(path, url_column)
    if suffix in (".xlsx", ".xlsm"):
        return _load_xlsx(path, url_column)
    raise ValueError(
        f"Formato no soportado: '{suffix}'. Soportados: .txt, .csv, .json, .xlsx"
    )


def _load_txt(path: Path) -> list[str]:
    """Una URL por línea. Ignora líneas vacías y comentarios (#)."""
    out: list[str] = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            stripped = line.strip()
            # Saltar vacías y comentarios — facilita anotar listas de auditoría.
            if not stripped or stripped.startswith("#"):
                continue
            out.append(stripped)
    return out


def _resolve_column(fieldnames: list[str], url_column: str | None) -> str:
    """Devuelve el nombre real de la columna URL o lanza ValueError descriptivo.

    Compara case-insensitive para tolerar cabeceras tipo "URL", "Url", "url".
    """
    lower_map = {f.lower(): f for f in fieldnames}
    if url_column is not None:
        real = lower_map.get(url_column.lower())
        if real is None:
            raise ValueError(
                f"La columna '{url_column}' no existe. "
                f"Columnas disponibles: {fieldnames}"
            )
        return real
    for candidate in DEFAULT_URL_COLUMNS:
        if candidate.lower() in lower_map:
            return lower_map[candidate.lower()]
    # Auto-detect falló: si hay columnas que contienen "url" en su nombre,
    # se las ofrecemos al usuario como sugerencia para --url-column.
    candidates = [f for f in fieldnames if "url" in f.lower()]
    if candidates:
        others = [f for f in fieldnames if f not in candidates]
        raise ValueError(
            "No se detectó columna URL automáticamente.\n"
            f'  Columnas con "url" en el nombre: {candidates}\n'
            f"  Otras columnas: {others}\n"
            f'  Usa --url-column "{candidates[0]}" para indicar cuál.'
        )
    raise ValueError(
        f"No se detectó columna URL. Columnas disponibles: {fieldnames}. "
        f"Usa --url-column para indicar cuál."
    )


def _load_csv(path: Path, url_column: str | None) -> list[str]:
    """Lee CSV con DictReader. Detecta o usa la columna indicada."""
    # utf-8-sig: silencia el BOM que dejan Excel y otros editores en Windows
    # (si lo hay) sin afectar a archivos que no lo tengan.
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise ValueError(f"CSV sin cabecera: {path}")
        column = _resolve_column(list(reader.fieldnames), url_column)
        return [row[column].strip() for row in reader if row.get(column, "").strip()]


def _load_xlsx(path: Path, url_column: str | None) -> list[str]:
    """Lee la primera hoja de un .xlsx/.xlsm con la misma lógica de columnas que CSV.

    - read_only + data_only: rápido en archivos grandes y devuelve el valor
      calculado de las fórmulas (no la fórmula en sí).
    - Convertimos cada celda a str: Excel guarda enteros como int y nadie
      espera "https://x.com/123" como `int("...")`.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = workbook.active  # primera hoja del workbook
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError(f"XLSX vacío: {path}")

        headers = [str(h) if h is not None else "" for h in header_row]
        column = _resolve_column(headers, url_column)
        column_index = headers.index(column)

        urls: list[str] = []
        for row in rows_iter:
            if column_index >= len(row):
                continue
            value = row[column_index]
            if value is None:
                continue
            text = str(value).strip()
            if text:
                urls.append(text)
        return urls
    finally:
        workbook.close()


def _load_json(path: Path, url_column: str | None) -> list[str]:
    """Soporta array de strings o array de objetos con campo URL detectable."""
    with path.open(encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("El JSON debe ser una lista de strings o de objetos.")
    if not data:
        return []
    first = data[0]
    if isinstance(first, str):
        return [item for item in data if isinstance(item, str)]
    if isinstance(first, dict):
        column = _resolve_column(list(first.keys()), url_column)
        return [
            item[column]
            for item in data
            if isinstance(item, dict) and column in item
        ]
    raise ValueError(
        f"Tipo de elemento JSON no soportado: {type(first).__name__}"
    )
